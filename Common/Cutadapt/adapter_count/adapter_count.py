from pathlib import Path
import subprocess
import re
import pandas as pd
from collections import OrderedDict

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook
import copy
from natsort import natsort_keygen
##########################################  excel ###########################################

def styles_copy(source_cell, target_cell):
    target_cell._style = copy.copy(source_cell._style)
    target_cell.font = copy.copy(source_cell.font)
    target_cell.border = copy.copy(source_cell.border)
    target_cell.fill = copy.copy(source_cell.fill)
    target_cell.number_format = copy.copy(source_cell.number_format)
    target_cell.protection = copy.copy(source_cell.protection)
    target_cell.alignment = copy.copy(source_cell.alignment)
    return target_cell


def copy_sheet_all(sheet, sheet2):
    """
    复制sheet到sheet2中，带格式复制
    """
    # tab颜色
    sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

    # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
    wm = list(sheet.merged_cells)
    if len(wm) > 0:
        for i in range(0, len(wm)):
            cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
            sheet2.merge_cells(cell2)

    for i, row in enumerate(sheet.iter_rows()):
        sheet2.row_dimensions[i+1].height = sheet.row_dimensions[i+1].height
        for j, cell in enumerate(row):
            sheet2.column_dimensions[get_column_letter(j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width
            sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

            # 设置单元格格式
            source_cell = sheet.cell(i+1, j+1)
            target_cell = sheet2.cell(i+1, j+1)
            target_cell.fill = copy.copy(source_cell.fill)
            if source_cell.has_style:
                styles_copy(source_cell, target_cell)
    return sheet2



def create_copy_sheets(wb, wbx, exists_names=[], allsheet=True, sub_sheetlst=[]):
    """将wbx中的所有/部分sheet写入到wb
    wb yes: sheet插入目标工作簿
    wbx yes: 源工作簿
    exists_names: 目标工作簿中已经存在的sheet名称
    allsheet: 是否将wbx中的所有sheet写入到wb
    sub_sheetlst: 需要写入的sheet名称列表
    """
    for n in wbx.sheetnames:
        if not allsheet:
            if n not in sub_sheetlst:
                continue
        if n in exists_names:
            sheetname = n+'_a'  # 同名sheet, 改为sheet_name+'_a'
        else:
            sheetname = n
        ws = wb.create_sheet(title=sheetname)
        # 不带格式赋值 sheet
        # wsc = copy_sheet(wbx[n], ws)
        # 带格式赋值 sheet
        wsc = copy_sheet_all(wbx[n], ws)
    return wb


##########################################  excel ###########################################

# https://www.cnblogs.com/hanfe1/p/12885200.html

#process = subprocess.Popen(["ls","-l"], shell=False,stdout=subprocess.PIPE, stderr=subprocess.PIPE,text=True)

# 正则匹配多次
# Sequence: CACCCTCTATCTCGAGAAAGCTCC; Type: regular 3'; Length: 24; Trimmed: 12318 times; Reverse-complemented: 51 times
# Sequence: ACATCTATATCACTATCCCGAACC; Type: regular 3'; Length: 24; Trimmed: 1 times; Reverse-complemented: 0 times


# '''
# args：需要执行的系统命令，可为字符串序列（列表或元组，shell为默认值False即可，建议为序列），也可为字符串（使用字符串时，需将shell赋值为True）；
# shell：默认为False，若args为序列时，shell=False；若args为字符串时，shell=True，表示通过shell执行命令；
# bufsize：指定缓冲策略，0表示不缓冲，1表示行缓冲，其它整数表示缓冲区大小，负数表示使用系统默认值0；
# cwd：默认值为None；若非None，则表示将会在执行这个子进程之前改变当前工作目录；
# env：用于指定子进程的环境变量。若env为None，那么子进程的环境变量将从父进程中继承；若env非None，则表示子程序的环境变量由env值来设置，它的值必须是一个映射对象。
# universal_newlines： 不同系统的换行符不同。若True，则该文件对象的stdin，stdout和stderr将会以文本流方式打开；否则以二进制流方式打开。
# '''

# #cwd="/path",  #指定运行目录
# # env={"key":"value"}  #指定环境变量
# # timeout=10,  #指定超时时间 秒(S)
# # universal_newlines=True  #指定输出为文本模式
# # stdout=subprocess.PIPE, stderr=subprocess.PIPE  #指定输出和错误输出


# # 读取标准输出和标准错误
# stdout, stderr = process.communicate()
# print(stdout)
# print(stderr+'#')
# # 返回值
# print(process.returncode)# !=0


def subprocess_run(adapter_lst,fastq_path,work_directory,adapter_type='adapter'):
    if adapter_type == 'adapter': #3'
        adapter_srt = " --adapter " + " --adapter ".join(adapter_lst)
    elif adapter_type == 'front': #5'
        adapter_srt = " --front " + " --front ".join(adapter_lst)
    elif adapter_type == 'anywhere': #5' and 3'
        adapter_srt = " --anywhere " + " --anywhere ".join(adapter_lst)
    process = subprocess.Popen('''
    /data/home/liuyinzhe/software/miniconda3/bin/cutadapt  \
    {adapter_srt} \
    --error-rate 0 \
    --overlap  9 \
    --no-indels \
    --revcomp \
    --action lowercase  \
    --cores 8 \
    {fastq_path}  >/dev/null
    '''.format(adapter_srt=adapter_srt,fastq_path=fastq_path)
    , shell=True,stdout=subprocess.PIPE, stderr=subprocess.PIPE,text=True,cwd=work_directory)

    # 读取标准输出和标准错误
    stdout, stderr = process.communicate()
    return stdout, stderr


def parse_stderr(stderr):
    '''
    # === Adapter 1 ===

    # Sequence: GCATGCCCTGCCCCTAAGAATTCG; Type: regular 3'; Length: 24; Trimmed: 118605 times; Reverse-complemented: 308 times

    # === Adapter 2 ===

    # Sequence: ACATCTATATCACTATCCCGAACC; Type: regular 3'; Length: 24; Trimmed: 1 times; Reverse-complemented: 0 times
    # === Summary ===

    # Total reads processed:                 329,176
    # Reads with adapters:                   118,606 (36.0%)
    # Reverse-complemented:                      308 (0.1%)

    # match_obj_lst = re.search(r'Sequence: (.*); Type: regular 3\'; Length: (\d+); Trimmed: (\d+) times; Reverse-complemented: (\d+) times', stdout)
    # #print(type(match_obj_lst))
    # print(match_obj_lst)

    '''
    total_read = re.search(r'Total reads processed:\s+([\d,]+)', stderr).group(1)
    total_read = int(re.sub(',','',total_read))
    total_trimmed_read = re.search(r'Reads with adapters:\s+([\d,]+) \([\d.]+%\)',stderr).group(1)
    total_trimmed_read= int(re.sub(',','',total_trimmed_read))
    reverse_complemented_read = re.search(r'Reverse-complemented:\s+([\d,]+) \([\d.]+%\)',stderr).group(1)
    reverse_complemented_read = int(re.sub(',','',reverse_complemented_read))
    
    #print(total_read,total_trimmed_read,reverse_complemented_read)
    match_info_lst = re.findall(r'Sequence: (.*); Type: regular 3\'; Length: (\d+); Trimmed: (\d+) times; Reverse-complemented: (\d+) times', stderr)
    return total_read,total_trimmed_read,reverse_complemented_read,match_info_lst

def GetAllFilePaths(pwd,wildcard='*'):
    '''
    获取目录下文件全路径，通配符检索特定文件名，返回列表
    param: str  "pwd"
    return:dirname pathlab_obj
    return:list [ str ]
    #https://zhuanlan.zhihu.com/p/36711862
    #https://www.cnblogs.com/sigai/p/8074329.html
    '''
    files_lst = []
    target_path=Path(pwd)
    for child in target_path.rglob(wildcard):
        if child.is_dir():
            pass
        elif child.is_file():
            files_lst.append(child)
    return files_lst



#  获得目录
current_dir = Path.cwd()
# 读取接头信息
adapter_tsv = current_dir.joinpath('adapter.tsv')

# 获取全部接头序列，以及序列名称
adapter_lst=[]
adapter_name_dic = {}
with open(adapter_tsv,'r') as fh:
    for line in fh:
        '''
        STM_5	TGCTAGAGGGCGGGAGAGTT
        '''
        record = re.split('\t',line.strip())
        adapter_name,adapter_seq = record
        if adapter_seq not in adapter_name_dic:
            adapter_name_dic[adapter_seq]=adapter_name
            adapter_lst.append(adapter_seq)
        else:
            raise Exception("重复的接头序列：!", adapter_seq)




# rawdata 目录
rawdata_dir = current_dir.joinpath('rawdata')
# 分析目录
statistics_directory = current_dir.joinpath('adapter_statistics')
statistics_directory.mkdir(exist_ok=True)

# 获取样品信息
sample_lst_file = current_dir.joinpath('sample.lst')
sample_name_lst = []
with open(sample_lst_file,'r') as fh:
    for line in fh:
        sample_name = line.strip()
        sample_name_lst.append(sample_name)

header_lst = ["样品名","总reads数量","含接头的reads数量","接头名字","接头序列","接头长度","包含该接头的reads数量","该接头在含接头的reads中占比(%)"] #"反向互补序列接头","反向互补接头百分占比"

for  sample_name in sample_name_lst:

    # 创建dataframe,用于后续保存结果
    #df = pd.DataFrame({'col1': [1, 2], 'col2': [3, 4]})  
    #header_lst = ["样品名","总reads数量","含接头的reads数量","接头名字","接头序列","接头长度","包含该接头的reads数量","该接头在含接头的reads中占比(%)"] #"反向互补序列接头","反向互补接头百分占比"
    header_dic = OrderedDict()
    for header_item in header_lst:
        header_dic[header_item] = []
    df = pd.DataFrame(header_dic)  

    # 获取 fastq文件路径
    fastq_path=rawdata_dir.joinpath(sample_name+'.1.fq.gz') # read1
    # 常见样品目录
    sample_dir=statistics_directory.joinpath(sample_name)
    sample_dir.mkdir(exist_ok=True)

    # 执行分析内容
    stdout, stderr = subprocess_run(adapter_lst,fastq_path,sample_dir)
    #print(stdout)

    # 获得stderr 中的接头信息
    total_read,total_trimmed_read,reverse_complemented_read,match_info_lst = parse_stderr(stderr)

    # 迭代 match_info_lst 向结果表格中记录
    for match_info in match_info_lst:
        adapter_seq, adapter_len, trimmed_times, reverse_complemented_times = match_info
        '''
        header_lst = ["样品名","总reads数量","含接头的reads数量","接头名字","接头序列","接头长度","包含该接头的reads数量","该接头在含接头的reads中占比"]
        '''
        # trimmed_times 当前接头trim的数量
        # reverse_complemented_times 当前接头trim 的数量
        adapter_name = adapter_name_dic[adapter_seq]
        trimmed_times = int(trimmed_times)
        trimmed_read_percent = round(trimmed_times/total_trimmed_read,4)*100
        #rc_read_percent = round(reverse_complemented_read/total_trimmed_read,4)*100
        add_lst = [sample_name,total_read,total_trimmed_read,adapter_name,adapter_seq,adapter_len,trimmed_times,trimmed_read_percent]
        # 逐行添加
        #df = df.append(pd.DataFrame([add_lst], columns=df.columns),ignore_index=True) # 旧方法
        df = pd.concat([df,pd.DataFrame([add_lst], columns=df.columns)], ignore_index=True)
    sample_xlsx = sample_dir.joinpath(sample_name+'.adapter_statistics.xlsx')
    df.to_excel(sample_xlsx, index=False,sheet_name=sample_name)


files_lst = GetAllFilePaths(statistics_directory,wildcard='*.adapter_statistics.xlsx')

# 创建dataframe,用于后续保存结果 

header_dic = OrderedDict()
for header_item in header_lst:
    header_dic[header_item] = []
df = pd.DataFrame(header_dic)

for xlsx_file in files_lst:
    item_df= pd.read_excel(xlsx_file)
    #df = df.append(item_df,ignore_index=True)
    df = pd.concat([df,item_df], ignore_index=True)
merged_xlsx = statistics_directory.joinpath('adapter_statistics.merged_type1.xlsx')
df.sort_values(by=['样品名','接头名字'],ascending=[True,True],inplace=True)
df.to_excel(merged_xlsx, index=False)


merged2_xlsx = statistics_directory.joinpath('adapter_statistics.merged_type2.xlsx')

# 合并第二种风格的文件

# 空的目标
wb = Workbook()
ex_names = []
sorted_files_lst = sorted(files_lst,key=lambda x:int(x.name.split(".")[0]))
for  xlsx_file in sorted_files_lst:
    wb_obj = load_workbook(xlsx_file)
    ex_names = wb.sheetnames
    #print(ex_names)
    if len(ex_names) == 0:
        wb = create_copy_sheets(wb, wb_obj, exists_names=[])
        # 已经有的sheet 名称
        ex_names = wb.sheetnames
    else:
        wb = create_copy_sheets(wb, wb_obj, exists_names=ex_names)  
# 保存前删除第一个空白Sheet
del wb["Sheet"] 
#保存合并后内容
wb.save(merged2_xlsx)
